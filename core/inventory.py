from openpyxl import load_workbook
from tabulate import tabulate
from config.paths import EXCEL_TEMPLATE
from config.settings import settings
from core.utils import format_number

class InventoryManager:
    def __init__(self, workbook):
        self.workbook = workbook
        self.sheet = workbook["Sales & Stocks"]
        self.data = self._load_data()
        self.stock, self.keys, self.categories, self.sizes = self._process_data()
        
    def _load_data(self):
        """Load and process inventory data from Excel"""
        data = []
        for row in range(2, self.sheet.max_row + 1):
            item = [
                self.sheet.cell(row=row, column=col).value 
                for col in range(1, 11)
            ]
            item[-2] = eval(item[-2][1:])  # Stock
            item[-1] = eval(item[-1][1:])  # Sale
            item.append(item[-2] - item[-1])  # Balance
            data.append(item)
        return data
    
    def _process_data(self):
        """Process data into dictionaries for quick access"""
        stock = {}
        categories = {}
        sizes = {}
        
        for item in self.data:
            code = item[2]
            stock[code] = item[3:] + item[:1]  # All details + base code
            
            # Categorize
            category = item[1]
            if category not in categories:
                categories[category] = []
            categories[category].append(item[2:])
            
            # Size mapping
            size = item[3]
            if size not in sizes:
                sizes[size] = [item[0], item[4]]  # Base code and MRP
            sizes[size].append(item[4])
            
        return stock, stock.keys(), categories, sizes
    
    def show_stock(self, category=None, code=None):
        """Display stock in table format"""
        headers = ["Code", "Name", "Size", "Unit", "MRP", "Price", "Stock", "Sale", "Balance"]
        
        if code:
            item = [[code] + self.stock.get(code)[:-1]]
            print(tabulate(item, headers=headers, tablefmt="fancy_grid"))
        elif category:
            print(tabulate(self.categories.get(category, []), headers=headers, tablefmt="fancy_grid"))
        else:
            print(tabulate([item[2:] for item in self.data], headers=headers, tablefmt="fancy_grid"))
    
    def add_stock(self, code, quantity):
        """Increase stock quantity for an item"""
        if code not in self.keys:
            print("Invalid item code!")
            return False
            
        stock_col = self._get_column_address("I")  # Stock column
        address = stock_col.get(code)
        current = self.sheet[address].value
        self.sheet[address] = current + f"+{quantity}"
        return True
    
    def add_item(self, details):
        """Add new item to inventory"""
        # Implementation omitted for brevity
        pass
    
    def _get_column_address(self, column_char):
        """Map item codes to column addresses"""
        return {item[2]: f"{column_char}{i+2}" for i, item in enumerate(self.data)}